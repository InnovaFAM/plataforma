from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from constants import EXPORT_COLLABS_COLUMNS
from utils import normalize_excel_value, unwrap_dynamodb_value


def build_collabs_excel(items: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Colaboradores"

    headers = [header for _, header in EXPORT_COLLABS_COLUMNS]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    column_widths = [len(header) for header in headers]

    for item in items:
        normalized_item = unwrap_dynamodb_value(item)

        row = []
        for index, (path, _) in enumerate(EXPORT_COLLABS_COLUMNS):
            value = resolve_export_value(normalized_item, path)
            value = normalize_excel_value(value)
            row.append(value)

            value_length = len(str(value)) if value is not None else 0
            column_widths[index] = max(column_widths[index], value_length)

        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for index, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(index)
        ws.column_dimensions[column_letter].width = min(max(width + 2, 12), 55)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def resolve_export_value(item: dict[str, Any], path: str) -> Any:
    if path == "assignments":
        return format_assignments(item.get("assignments"))

    value: Any = item

    for part in path.split("."):
        if not isinstance(value, dict):
            return None
        value = value.get(part)

    return value


def format_assignments(assignments: Any) -> str:
    assignments = unwrap_dynamodb_value(assignments)

    if not assignments:
        return ""

    if not isinstance(assignments, list):
        return str(assignments)

    service_names = []

    for assignment in assignments:
        assignment = unwrap_dynamodb_value(assignment)

        if isinstance(assignment, dict):
            service_name = assignment.get("serviceName")
            service_sk = assignment.get("sk")

            if service_name:
                service_names.append(str(service_name))
            elif service_sk:
                service_names.append(str(service_sk))
        else:
            service_names.append(str(assignment))

    return " | ".join(service_names)


def build_file_name(prefix: str) -> str:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.xlsx"
