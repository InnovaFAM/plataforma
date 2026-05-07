from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from constants import SERVICE_EXPORT_COLUMNS
from utils import normalize_excel_value, unwrap_dynamodb_value


def build_services_excel(items: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios"

    headers = [header for _, header in SERVICE_EXPORT_COLUMNS]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    column_widths = [len(header) for header in headers]

    for item in items:
        normalized_item = unwrap_dynamodb_value(item)

        row = []

        for index, (path, _) in enumerate(SERVICE_EXPORT_COLUMNS):
            value = resolve_service_export_value(normalized_item, path)
            value = normalize_excel_value(value)

            row.append(value)

            value_length = len(str(value)) if value is not None else 0
            column_widths[index] = max(column_widths[index], value_length)

        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for index, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(index)
        ws.column_dimensions[column_letter].width = min(max(width + 2, 12), 55)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def resolve_service_export_value(item: dict[str, Any], path: str) -> Any:
    if path == "managers":
        return format_people(item.get("managers"))

    if path == "submanagers":
        return format_people(item.get("submanagers"))

    value: Any = item

    for part in path.split("."):
        if not isinstance(value, dict):
            return None

        value = value.get(part)

    return value


def format_people(people: Any) -> str:
    people = unwrap_dynamodb_value(people)

    if not people:
        return ""

    if not isinstance(people, list):
        return str(people)

    formatted_people = []

    for person in people:
        person = unwrap_dynamodb_value(person)

        if not isinstance(person, dict):
            formatted_people.append(str(person))
            continue

        name = person.get("name")
        role = person.get("role")
        email = person.get("email")
        phone_number = person.get("phoneNumber")
        person_type = person.get("type")

        parts = []

        if name:
            parts.append(str(name))

        metadata = []

        if role:
            metadata.append(str(role))

        if person_type:
            metadata.append(str(person_type))

        if metadata:
            parts.append(f"({', '.join(metadata)})")

        contact = []

        if email:
            contact.append(str(email))

        if phone_number:
            contact.append(str(phone_number))

        if contact:
            parts.append("- " + " / ".join(contact))

        formatted_people.append(" ".join(parts))

    return " | ".join(formatted_people)
