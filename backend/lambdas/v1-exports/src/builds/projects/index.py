from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from builds.projects.utils import add_excel_table, translate_role_status
from utils import (
    apply_table_style,
    autosize_columns,
    normalize_excel_value,
    unwrap_dynamodb_value,
)


def build_projects_panel_excel(report: dict[str, Any]) -> bytes:
    report = unwrap_dynamodb_value(report)

    # Explicitly ignore filters
    report.pop("filters", None)

    wb = Workbook()

    ws_dashboard = wb.active
    ws_dashboard.title = "Resumen"

    ws_population = create_population_by_project_sheet(wb, report)
    ws_schedule = create_service_schedule_sheet(wb, report)
    ws_services = create_staffing_by_service_sheet(wb, report)
    ws_roles = create_roles_table_sheet(wb, report)

    create_projects_dashboard_sheet(
        ws=ws_dashboard,
        report=report,
        population_ws=ws_population,
        schedule_ws=ws_schedule,
    )

    output = BytesIO()
    wb.save(output)

    return output.getvalue()


def create_projects_dashboard_sheet(
    ws: Any,
    report: dict[str, Any],
    population_ws: Any,
    schedule_ws: Any,
) -> None:
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = "Panel de proyectos"
    title_cell.font = Font(size=18, bold=True, color="1F2937")
    title_cell.alignment = Alignment(vertical="center")

    ws.merge_cells("A2:N2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = "Visualización de dotación y recursos por servicio"
    subtitle_cell.font = Font(size=11, color="6B7280")

    kpis = report.get("kpis") or {}

    kpi_cards = [
        ("Requeridos", kpis.get("totalRequired", 0)),
        ("Confirmados", kpis.get("totalConfirmed", 0)),
        ("Propuestos", kpis.get("totalProposed", 0)),
        ("Dotación real", kpis.get("totalReal", 0)),
        ("Brecha", kpis.get("totalGap", 0)),
        ("Vacantes", kpis.get("totalMissing", 0)),
        ("Excedentes", kpis.get("totalSurplus", 0)),
    ]

    create_kpi_cards(ws, kpi_cards)

    add_population_chart_to_dashboard(
        dashboard_ws=ws,
        data_ws=population_ws,
        anchor="A9",
    )

    add_schedule_chart_to_dashboard(
        dashboard_ws=ws,
        data_ws=schedule_ws,
        anchor="J9",
    )

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 4
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 4
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 16
    ws.column_dimensions["M"].width = 16
    ws.column_dimensions["N"].width = 16

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22


def create_kpi_cards(
    ws: Any,
    kpi_cards: list[tuple[str, Any]],
) -> None:
    title_fill = PatternFill("solid", fgColor="F3F4F6")
    value_fill = PatternFill("solid", fgColor="FFFFFF")
    border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    start_row = 4

    for index, (label, value) in enumerate(kpi_cards):
        block_col = 1 + (index % 4) * 3
        block_row = start_row + (index // 4) * 3

        label_cell = ws.cell(row=block_row, column=block_col)
        value_cell = ws.cell(row=block_row + 1, column=block_col)

        ws.merge_cells(
            start_row=block_row,
            start_column=block_col,
            end_row=block_row,
            end_column=block_col + 1,
        )

        ws.merge_cells(
            start_row=block_row + 1,
            start_column=block_col,
            end_row=block_row + 1,
            end_column=block_col + 1,
        )

        label_cell.value = label
        label_cell.fill = title_fill
        label_cell.font = Font(size=10, bold=True, color="6B7280")
        label_cell.alignment = Alignment(horizontal="center", vertical="center")

        value_cell.value = normalize_excel_value(value)
        value_cell.fill = value_fill
        value_cell.font = Font(size=16, bold=True, color="111827")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in range(block_row, block_row + 2):
            for col in range(block_col, block_col + 2):
                ws.cell(row=row, column=col).border = border

        if (
            label in {"Brecha", "Vacantes"}
            and isinstance(value, (int, float))
            and value != 0
        ):
            value_cell.font = Font(size=16, bold=True, color="DC2626")


def create_population_by_project_sheet(
    wb: Workbook,
    report: dict[str, Any],
) -> Any:
    ws = wb.create_sheet("Poblamiento")

    rows = report.get("charts", {}).get("populationByProject", [])

    headers = [
        "Servicio ID",
        "Servicio",
        "Requeridos",
        "Confirmados",
        "Propuestos",
        "Real",
        "Brecha",
        "Faltantes",
        "Excedentes",
    ]

    ws.append(headers)

    for item in rows:
        ws.append(
            [
                item.get("serviceId"),
                item.get("serviceName"),
                item.get("required", 0),
                item.get("confirmed", 0),
                item.get("proposed", 0),
                item.get("real", 0),
                item.get("gap", 0),
                item.get("missingCount", 0),
                item.get("surplusCount", 0),
            ]
        )

    apply_table_style(ws)
    autosize_columns(ws)
    add_excel_table(ws, "TblPoblamiento")

    if ws.max_row >= 2:
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        ws.conditional_formatting.add(
            f"G2:G{ws.max_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=red_fill),
        )

    return ws


def add_population_chart_to_dashboard(
    dashboard_ws: Any,
    data_ws: Any,
    anchor: str,
) -> None:
    if data_ws.max_row < 2:
        return

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Poblamiento por proyectos"
    chart.y_axis.title = "Cantidad"
    chart.x_axis.title = "Servicio"
    chart.height = 9
    chart.width = 17

    data = Reference(
        data_ws,
        min_col=3,
        max_col=5,
        min_row=1,
        max_row=data_ws.max_row,
    )

    categories = Reference(
        data_ws,
        min_col=1,
        min_row=2,
        max_row=data_ws.max_row,
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    dashboard_ws.add_chart(chart, anchor)


def create_service_schedule_sheet(
    wb: Workbook,
    report: dict[str, Any],
) -> Any:
    ws = wb.create_sheet("Programación")

    rows = report.get("charts", {}).get("serviceSchedule", [])

    headers = [
        "Mes",
        "Requeridos",
        "Confirmados",
        "Propuestos",
        "Real",
        "Brecha",
        "Faltantes",
        "Excedentes",
    ]

    ws.append(headers)

    for item in rows:
        ws.append(
            [
                item.get("month"),
                item.get("required", 0),
                item.get("confirmed", 0),
                item.get("proposed", 0),
                item.get("real", 0),
                item.get("gap", 0),
                item.get("missingCount", 0),
                item.get("surplusCount", 0),
            ]
        )

    apply_table_style(ws)
    autosize_columns(ws)
    add_excel_table(ws, "TblProgramacion")

    if ws.max_row >= 2:
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        ws.conditional_formatting.add(
            f"F2:F{ws.max_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=red_fill),
        )

    return ws


def add_schedule_chart_to_dashboard(
    dashboard_ws: Any,
    data_ws: Any,
    anchor: str,
) -> None:
    if data_ws.max_row < 2:
        return

    chart = LineChart()
    chart.style = 13
    chart.title = "Programación de servicios"
    chart.y_axis.title = "Cantidad"
    chart.x_axis.title = "Mes"
    chart.height = 9
    chart.width = 17

    data = Reference(
        data_ws,
        min_col=2,
        max_col=4,
        min_row=1,
        max_row=data_ws.max_row,
    )

    categories = Reference(
        data_ws,
        min_col=1,
        min_row=2,
        max_row=data_ws.max_row,
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    dashboard_ws.add_chart(chart, anchor)


def create_staffing_by_service_sheet(
    wb: Workbook,
    report: dict[str, Any],
) -> Any:
    ws = wb.create_sheet("Dotación Servicio")

    population = report.get("charts", {}).get("populationByProject", [])
    roles = report.get("rolesTable", [])

    role_counts: dict[str, int] = {}

    for role in roles:
        service_id = role.get("serviceId")

        if not service_id:
            continue

        role_counts[service_id] = role_counts.get(service_id, 0) + 1

    headers = [
        "Servicio ID",
        "Servicio",
        "Roles",
        "Requeridos",
        "Confirmados",
        "Propuestos",
        "Real",
        "Brecha",
        "Vacantes",
        "Excedentes",
    ]

    ws.append(headers)

    for item in population:
        service_id = item.get("serviceId")

        ws.append(
            [
                service_id,
                item.get("serviceName"),
                role_counts.get(service_id, 0),
                item.get("required", 0),
                item.get("confirmed", 0),
                item.get("proposed", 0),
                item.get("real", 0),
                item.get("gap", 0),
                item.get("missingCount", 0),
                item.get("surplusCount", 0),
            ]
        )

    apply_table_style(ws)
    autosize_columns(ws)
    add_excel_table(ws, "TblDotacionServicio")

    if ws.max_row >= 2:
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        ws.conditional_formatting.add(
            f"H2:H{ws.max_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=red_fill),
        )

    return ws


def create_roles_table_sheet(
    wb: Workbook,
    report: dict[str, Any],
) -> Any:
    ws = wb.create_sheet("Detalle Roles")

    rows = report.get("rolesTable", [])

    headers = [
        "Servicio ID",
        "Servicio",
        "Rol ID",
        "Rol SK",
        "Rol",
        "Requeridos",
        "Confirmados",
        "Propuestos",
        "Real",
        "Brecha",
        "Vacantes",
        "Excedentes",
        "Estado",
        "Inicio",
        "Fin",
    ]

    ws.append(headers)

    for item in rows:
        ws.append(
            [
                item.get("serviceId"),
                item.get("serviceName"),
                item.get("roleId"),
                item.get("roleSk"),
                item.get("roleName"),
                item.get("requiredCount", 0),
                item.get("confirmedCount", 0),
                item.get("proposedCount", 0),
                item.get("realCount", 0),
                item.get("gap", 0),
                item.get("missingCount", 0),
                item.get("surplusCount", 0),
                translate_role_status(item.get("status")),
                item.get("startedAt"),
                item.get("endedAt"),
            ]
        )

    apply_table_style(ws)
    autosize_columns(ws)
    add_excel_table(ws, "TblDetalleRoles")
    apply_roles_conditional_formatting(ws)

    return ws


def apply_roles_conditional_formatting(ws: Any) -> None:
    if ws.max_row < 2:
        return

    red_fill = PatternFill("solid", fgColor="FEE2E2")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    yellow_fill = PatternFill("solid", fgColor="FEF3C7")

    # Brecha
    ws.conditional_formatting.add(
        f"J2:J{ws.max_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill),
    )

    # Estado
    for row in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=row, column=13)

        if status_cell.value == "Faltante":
            status_cell.fill = red_fill
            status_cell.font = Font(color="DC2626", bold=True)

        elif status_cell.value == "Completo":
            status_cell.fill = green_fill
            status_cell.font = Font(color="166534", bold=True)

        elif status_cell.value == "Excedente":
            status_cell.fill = yellow_fill
            status_cell.font = Font(color="92400E", bold=True)
