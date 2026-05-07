from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def translate_role_status(status: str | None) -> str:
    normalized = (status or "").lower()

    if normalized == "missing":
        return "Faltante"

    if normalized == "complete":
        return "Completo"

    if normalized == "surplus":
        return "Excedente"

    return status or ""


def add_excel_table(ws: Any, table_name: str) -> None:
    if ws.max_row < 2:
        return

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    table = Table(
        displayName=table_name,
        ref=ref,
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = style
    ws.add_table(table)
