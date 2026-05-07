from calendar import monthrange
from datetime import date, datetime, timezone
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def resolve_timeline_bounds(
    service: dict[str, Any],
    roles: list[dict[str, Any]],
    assignments_by_role: dict[str, list[dict[str, Any]]],
) -> tuple[date, date]:
    dates: list[date] = []

    for value in [service.get("startDate"), service.get("endDate")]:
        parsed = parse_date(value)
        if parsed:
            dates.append(parsed)

    for role in roles:
        for value in [role.get("startedAt"), role.get("endedAt")]:
            parsed = parse_date(value)
            if parsed:
                dates.append(parsed)

        role_sk = role.get("sk")

        for assignment in assignments_by_role.get(role_sk, []):
            for value in [assignment.get("startedAt"), assignment.get("endedAt")]:
                parsed = parse_date(value)
                if parsed:
                    dates.append(parsed)

    if not dates:
        today = datetime.now(timezone.utc).date()
        return today, today

    return min(dates), max(dates)


def build_month_timeline(
    start_date: date,
    end_date: date,
) -> list[date]:
    start = date(start_date.year, start_date.month, 1)
    end = date(end_date.year, end_date.month, 1)

    months: list[date] = []
    current = start

    while current <= end:
        months.append(current)

        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)

    return months


def paint_timeline_cells(
    ws,
    row_number: int,
    months: list[date],
    start_date: date | None,
    end_date: date | None,
    start_col: int,
    fill: PatternFill,
) -> None:
    if not start_date or not end_date:
        return

    for index, month_start in enumerate(months):
        month_end = get_month_end(month_start)

        if date_ranges_overlap(
            start_a=start_date,
            end_a=end_date,
            start_b=month_start,
            end_b=month_end,
        ):
            cell = ws.cell(row=row_number, column=start_col + index)
            cell.fill = fill
            cell.value = "●"
            cell.alignment = Alignment(horizontal="center", vertical="center")


def date_ranges_overlap(
    start_a: date,
    end_a: date,
    start_b: date,
    end_b: date,
) -> bool:
    return start_a <= end_b and end_a >= start_b


def get_month_end(month_start: date) -> date:
    last_day = monthrange(month_start.year, month_start.month)[1]
    return date(month_start.year, month_start.month, last_day)


def calculate_duration_days(
    start_date: date | None,
    end_date: date | None,
) -> int | str:
    if not start_date or not end_date:
        return ""

    return (end_date - start_date).days + 1


def parse_date(value: Any) -> date | None:
    if not value:
        return None

    if isinstance(value, date):
        return value

    if not isinstance(value, str):
        return None

    try:
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def format_date(value: date | None) -> str:
    if not value:
        return ""

    return value.strftime("%Y-%m-%d")


def resolve_assignment_fill(
    status: str,
    proposed_fill: PatternFill,
    confirmed_fill: PatternFill,
    default_fill: PatternFill,
    inactive_fill: PatternFill,
) -> PatternFill:
    normalized = (status or "").lower()

    if normalized in {"confirmado", "confirmed"}:
        return confirmed_fill

    if normalized in {"propuesto", "proposed"}:
        return proposed_fill

    if normalized in {"inactivo", "inactive", "terminado"}:
        return inactive_fill

    return default_fill


def get_nested(item: dict[str, Any], path: str) -> Any:
    value: Any = item

    for part in path.split("."):
        if not isinstance(value, dict):
            return None

        value = value.get(part)

    return value
