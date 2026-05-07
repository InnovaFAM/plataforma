from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from builds.service_detail.utils import (
    build_month_timeline,
    calculate_duration_days,
    format_date,
    get_nested,
    paint_timeline_cells,
    parse_date,
    resolve_assignment_fill,
    resolve_timeline_bounds,
)
from builds.services.index import format_people
from constants import SERVICE_DETAIL_ASSIGNMENT_COLUMNS, SERVICE_DETAIL_ROLE_COLUMNS
from utils import (
    apply_borders,
    apply_table_style,
    autosize_columns,
    normalize_excel_value,
)


def build_service_detail_excel(detail: dict[str, Any]) -> bytes:
    wb = Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    create_service_sheet(wb, detail["service"])
    create_roles_sheet(wb, detail["roles"], detail["assignmentsByRole"])
    create_assignments_sheet(wb, detail["assignments"])
    create_gantt_sheet(
        wb,
        service=detail["service"],
        roles=detail["roles"],
        assignments_by_role=detail["assignmentsByRole"],
    )

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def create_service_sheet(
    wb: Workbook,
    service: dict[str, Any],
) -> None:
    ws = wb.create_sheet("Servicio")

    rows = [
        ("Código", service.get("code")),
        ("Nombre", service.get("name")),
        ("N° contrato", service.get("contractNumber")),
        ("Estado", service.get("status")),
        ("Prioridad", service.get("priority")),
        ("Fecha inicio", service.get("startDate")),
        ("Fecha término", service.get("endDate")),
        ("Cliente", get_nested(service, "client.name")),
        ("RUT cliente", get_nested(service, "client.rut")),
        ("SK cliente", get_nested(service, "client.sk")),
        ("División", get_nested(service, "division.name")),
        ("N° división", get_nested(service, "division.number")),
        ("SK división", get_nested(service, "division.sk")),
        ("Faena", get_nested(service, "chore.name")),
        ("Código faena", get_nested(service, "chore.code")),
        ("SK faena", get_nested(service, "chore.sk")),
        ("Administradores", format_people(service.get("managers"))),
        ("Subadministradores", format_people(service.get("submanagers"))),
        ("Parent ID", service.get("parentId")),
        ("SK", service.get("sk")),
    ]

    ws.append(["Campo", "Valor"])

    for label, value in rows:
        ws.append([label, normalize_excel_value(value)])

    apply_table_style(ws)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")


def create_roles_sheet(
    wb: Workbook,
    roles: list[dict[str, Any]],
    assignments_by_role: dict[str, list[dict[str, Any]]],
) -> None:
    ws = wb.create_sheet("Roles")

    headers = [header for _, header in SERVICE_DETAIL_ROLE_COLUMNS]
    ws.append(headers)

    for role in roles:
        role_sk = role.get("sk")
        assignments_count = len(assignments_by_role.get(role_sk, []))

        row = []

        for path, _ in SERVICE_DETAIL_ROLE_COLUMNS:
            if path == "assignmentsCount":
                value = assignments_count
            else:
                value = get_nested(role, path)

            row.append(normalize_excel_value(value))

        ws.append(row)

    apply_table_style(ws)
    autosize_columns(ws)


def create_assignments_sheet(
    wb: Workbook,
    assignments: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Asignaciones")

    headers = [header for _, header in SERVICE_DETAIL_ASSIGNMENT_COLUMNS]
    ws.append(headers)

    for assignment in assignments:
        row = []

        for path, _ in SERVICE_DETAIL_ASSIGNMENT_COLUMNS:
            value = get_nested(assignment, path)
            row.append(normalize_excel_value(value))

        ws.append(row)

    apply_table_style(ws)
    autosize_columns(ws)


def create_gantt_sheet(
    wb: Workbook,
    service: dict[str, Any],
    roles: list[dict[str, Any]],
    assignments_by_role: dict[str, list[dict[str, Any]]],
) -> None:
    ws = wb.create_sheet("Gantt")

    timeline_start, timeline_end = resolve_timeline_bounds(
        service=service,
        roles=roles,
        assignments_by_role=assignments_by_role,
    )

    months = build_month_timeline(timeline_start, timeline_end)

    fixed_headers = [
        "Tipo",
        "Rol",
        "Colaborador",
        "Estado",
        "Inicio",
        "Fin",
        "Duración días",
    ]

    timeline_headers = [month.strftime("%Y-%m") for month in months]

    ws.append(fixed_headers + timeline_headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    role_fill = PatternFill("solid", fgColor="D9EAF7")
    assignment_fill = PatternFill("solid", fgColor="C6E0B4")
    proposed_fill = PatternFill("solid", fgColor="FFF2CC")
    confirmed_fill = PatternFill("solid", fgColor="A9D18E")
    inactive_fill = PatternFill("solid", fgColor="E7E6E6")

    current_row = 2

    for role in roles:
        role_sk = role.get("sk")
        role_name = role.get("roleName", "")
        role_start = parse_date(role.get("startedAt"))
        role_end = parse_date(role.get("endedAt"))

        duration_days = calculate_duration_days(role_start, role_end)

        ws.append(
            [
                "Rol",
                role_name,
                "",
                "",
                format_date(role_start),
                format_date(role_end),
                duration_days,
            ]
        )

        for cell in ws[current_row]:
            cell.fill = role_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")

        paint_timeline_cells(
            ws=ws,
            row_number=current_row,
            months=months,
            start_date=role_start,
            end_date=role_end,
            start_col=len(fixed_headers) + 1,
            fill=role_fill,
        )

        current_row += 1

        assignments = assignments_by_role.get(role_sk, [])

        for assignment in assignments:
            assignment_start = parse_date(assignment.get("startedAt"))
            assignment_end = parse_date(assignment.get("endedAt"))
            assignment_status = assignment.get("status", "")
            collab_name = get_nested(assignment, "collab.name") or ""

            assignment_duration_days = calculate_duration_days(
                assignment_start,
                assignment_end,
            )

            ws.append(
                [
                    "Asignación",
                    role_name,
                    collab_name,
                    assignment_status,
                    format_date(assignment_start),
                    format_date(assignment_end),
                    assignment_duration_days,
                ]
            )

            fill = resolve_assignment_fill(
                status=assignment_status,
                proposed_fill=proposed_fill,
                confirmed_fill=confirmed_fill,
                default_fill=assignment_fill,
                inactive_fill=inactive_fill,
            )

            paint_timeline_cells(
                ws=ws,
                row_number=current_row,
                months=months,
                start_date=assignment_start,
                end_date=assignment_end,
                start_col=len(fixed_headers) + 1,
                fill=fill,
            )

            current_row += 1

        current_row += 1

    ws.freeze_panes = "H2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 16,
        "B": 34,
        "C": 36,
        "D": 18,
        "E": 14,
        "F": 14,
        "G": 16,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for col_index in range(
        len(fixed_headers) + 1, len(fixed_headers) + len(months) + 1
    ):
        column_letter = get_column_letter(col_index)
        ws.column_dimensions[column_letter].width = 12

    apply_borders(ws)
