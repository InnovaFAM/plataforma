import json
import uuid
from decimal import Decimal
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def unwrap_dynamodb_value(value: Any) -> Any:
    """
    Supports both native boto3 DynamoDB resource format and low-level
    DynamoDB AttributeValue format like {"S": "..."} or {"M": {...}}.
    """

    if isinstance(value, list):
        return [unwrap_dynamodb_value(item) for item in value]

    if not isinstance(value, dict):
        return value

    if len(value) == 1:
        key = next(iter(value.keys()))
        raw = value[key]

        if key == "S":
            return raw
        if key == "N":
            return Decimal(str(raw))
        if key == "BOOL":
            return bool(raw)
        if key == "NULL":
            return None
        if key == "M":
            return unwrap_dynamodb_value(raw)
        if key == "L":
            return [unwrap_dynamodb_value(item) for item in raw]
        if key in {"SS", "NS"}:
            return list(raw)

    return {k: unwrap_dynamodb_value(v) for k, v in value.items()}


def normalize_excel_value(value: Any) -> Any:
    value = unwrap_dynamodb_value(value)

    if isinstance(value, Decimal):
        if value % 1 == 0:
            return int(value)
        return float(value)

    if isinstance(value, bool):
        return "Sí" if value else "No"

    if value is None:
        return ""

    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)

    return value


def build_s3_key(to_email: str, folder: str, file_name: str) -> str:
    safe_email = (
        to_email.lower().replace("@", "_at_").replace(".", "_").replace("+", "_")
    )

    return f"exports/{folder}/{safe_email}/{uuid.uuid4()}_{file_name}"


def apply_table_style(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    apply_borders(ws)


def apply_borders(ws) -> None:
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 55)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
